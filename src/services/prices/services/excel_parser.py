"""Parser del Excel ``Lista de Precios Apro``.

Reglas observadas en la planilla actual (mayo 2026):

- Hoja única: ``LISTA DE PRECIOS`` (en mayúsculas; comparamos sin acentos).
- Fila 7: título.
- Fila 8: encabezados.
- Fila 9 en adelante: datos.
- Columnas (1-indexadas):

  | Col | Campo |
  |-----|-------|
  | A 1 | SAP |
  | B 2 | Modelo |
  | C 3 | Descripción |
  | D 4 | Familia |
  | E 5 | Categoría |
  | F 6 | Subcategoría |
  | G 7 | Marca |
  | H 8 | B2C/WEB (IVA incluido) |
  | I 9 | #1 PYME (valor Neto) |
  | J 10 | #2 Empresa-Distribuidor (valor Neto) |
  | K 11 | #3 Gran Empresa (valor Neto) |

El parser es **tolerante a duplicados de SAP**: gana la última fila (acordado
con el equipo de operaciones). Si el header no aparece, lanza ``ValueError``
con un mensaje accionable para el admin.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Iterable


HEADER_VARIANTS = {
    "sap": {"sap", "codigo sap", "cod sap", "codigo"},
    "modelo": {"modelo"},
    "descripcion": {"descripcion", "descripción"},
    "familia": {"familia"},
    "categoria": {"categoria", "categoría"},
    "subcategoria": {"subcategoria", "subcategoría"},
    "marca": {"marca"},
    "price_b2c_iva": {
        "b2c web iva incluido",
        "b2c web (iva incluido)",
        "b2c/web (iva incluido)",
        "b2c web",
        "b2c",
    },
    "price_pyme_neto": {
        "1 pyme valor neto",
        "1 pyme (valor neto)",
        "#1 pyme (valor neto)",
        "#1 pyme valor neto",
        "pyme",
        "pyme valor neto",
    },
    "price_distribuidor_neto": {
        "2 empresa distribuidor valor neto",
        "2 empresa-distribuidor (valor neto)",
        "#2 empresa distribuidor valor neto",
        "#2 empresa -distribuidor (valor neto)",
        "#2 empresa-distribuidor (valor neto)",
        "empresa distribuidor",
        "empresa-distribuidor",
        "mediana",
    },
    "price_gran_empresa_neto": {
        "3 gran empresa valor neto",
        "3 gran empresa (valor neto)",
        "#3 gran empresa (valor neto)",
        "#3 gran empresa valor neto",
        "gran empresa",
        "gran empresa valor neto",
    },
}

# Columnas obligatorias (sin estas no podemos persistir nada útil).
REQUIRED_HEADERS = {"sap"}

# Columnas con precios — son las útiles para downstream.
PRICE_FIELDS = (
    "price_b2c_iva",
    "price_pyme_neto",
    "price_distribuidor_neto",
    "price_gran_empresa_neto",
)


@dataclass(frozen=True)
class ParsedItem:
    sap_code: str
    modelo: str | None
    descripcion: str | None
    familia: str | None
    categoria: str | None
    subcategoria: str | None
    marca: str | None
    price_b2c_iva: Decimal | None
    price_pyme_neto: Decimal | None
    price_distribuidor_neto: Decimal | None
    price_gran_empresa_neto: Decimal | None

    def to_row_dict(self) -> dict[str, Any]:
        return {
            "sap_code": self.sap_code,
            "modelo": self.modelo,
            "descripcion": self.descripcion,
            "familia": self.familia,
            "categoria": self.categoria,
            "subcategoria": self.subcategoria,
            "marca": self.marca,
            "price_b2c_iva": self.price_b2c_iva,
            "price_pyme_neto": self.price_pyme_neto,
            "price_distribuidor_neto": self.price_distribuidor_neto,
            "price_gran_empresa_neto": self.price_gran_empresa_neto,
        }


@dataclass(frozen=True)
class ParseResult:
    items: list[ParsedItem]
    duplicates_overwritten: int
    rows_skipped: int


def _normalize_header(s: Any) -> str:
    raw = str(s or "").strip().lower()
    raw = unicodedata.normalize("NFD", raw)
    raw = "".join(c for c in raw if unicodedata.category(c) != "Mn")
    raw = re.sub(r"[^a-z0-9]+", " ", raw)
    return re.sub(r"\s+", " ", raw).strip()


def _match_header(value: Any) -> str | None:
    norm = _normalize_header(value)
    if not norm:
        return None
    for field, variants in HEADER_VARIANTS.items():
        if norm in variants:
            return field
    return None


def _coerce_text(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _coerce_sap(value: Any) -> str | None:
    """SAP puede venir como número largo o texto. Normalizamos a string sin
    espacios; si el origen es float (Excel mete ints como float) y es entero,
    lo convertimos a su representación entera para no perder ceros a la izq.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value).strip()
    return s or None


def _coerce_decimal(value: Any) -> Decimal | None:
    """Convierte un valor de celda en ``Decimal`` o ``None``.

    El Excel del cotizador suele venir con valores numéricos (openpyxl los
    entrega como ``int``/``float``). Si llega como string, primero intentamos
    el formato técnico ``187836.975`` y, si falla, asumimos formato local
    chileno ``187.836,975`` (punto miles, coma decimal).
    """
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return Decimal(s)
        except InvalidOperation:
            pass
        # Formato chileno: punto miles + coma decimal.
        s2 = s.replace(".", "").replace(",", ".")
        try:
            return Decimal(s2)
        except InvalidOperation:
            return None
    return None


def _find_header_row(ws, max_search_rows: int = 20) -> tuple[int, dict[str, int]]:
    """Localiza la fila del encabezado escaneando las primeras N filas.

    Devuelve ``(row_index, column_map)`` donde ``column_map[field] = col_idx``.
    """
    last_col = ws.max_column or 0
    if last_col == 0:
        raise ValueError("Hoja vacía: no hay columnas")

    upper = min(max_search_rows, ws.max_row or max_search_rows)
    best_row = 0
    best_map: dict[str, int] = {}
    for row in range(1, upper + 1):
        col_map: dict[str, int] = {}
        for col in range(1, last_col + 1):
            field = _match_header(ws.cell(row=row, column=col).value)
            if field and field not in col_map:
                col_map[field] = col
        if "sap" in col_map and any(p in col_map for p in PRICE_FIELDS):
            if len(col_map) > len(best_map):
                best_row = row
                best_map = col_map

    if not best_row or not REQUIRED_HEADERS.issubset(best_map.keys()):
        raise ValueError(
            "No se encontró el encabezado de la lista (esperaba una fila con "
            "'SAP' y al menos una columna de precio: PYME / Distribuidor / Gran Empresa)."
        )
    return best_row, best_map


def parse_price_list_excel(
    file_bytes: bytes,
    *,
    sheet_name: str | None = None,
) -> ParseResult:
    """Parsea el Excel de listas de precios y devuelve filas listas para insertar.

    - Si ``sheet_name`` se omite usa la primera hoja.
    - Si llega vacío o no tiene encabezado válido, lanza ``ValueError``.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "openpyxl no está instalado en la Lambda. Verificar requirements.txt"
        ) from e

    if not file_bytes:
        raise ValueError("El archivo está vacío")

    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_bytes), data_only=True, read_only=True
        )
    except Exception as e:
        raise ValueError(f"No se pudo abrir el Excel: {e}") from e

    target = sheet_name or wb.sheetnames[0]
    if target not in wb.sheetnames:
        # Match case-insensitive sin acentos
        norm_target = _normalize_header(target)
        match = next(
            (s for s in wb.sheetnames if _normalize_header(s) == norm_target),
            None,
        )
        if match is None:
            raise ValueError(
                f"Hoja '{target}' no existe. Hojas disponibles: {wb.sheetnames}"
            )
        target = match
    ws = wb[target]

    header_row, col_map = _find_header_row(ws)

    items: dict[str, ParsedItem] = {}
    duplicates = 0
    skipped = 0
    sap_col = col_map["sap"]

    rows_iter: Iterable = ws.iter_rows(
        min_row=header_row + 1, values_only=True
    )

    for raw_row in rows_iter:
        if raw_row is None:
            continue
        # ``values_only=True`` da una tupla 0-indexada de la fila completa.
        sap_val = raw_row[sap_col - 1] if sap_col - 1 < len(raw_row) else None
        sap_code = _coerce_sap(sap_val)
        if not sap_code:
            skipped += 1
            continue

        def _cell(field: str) -> Any:
            idx = col_map.get(field)
            if idx is None:
                return None
            i = idx - 1
            if i < 0 or i >= len(raw_row):
                return None
            return raw_row[i]

        prices = {
            f: _coerce_decimal(_cell(f)) for f in PRICE_FIELDS
        }
        if all(v is None for v in prices.values()):
            # Fila sin precios útiles: no la guardamos.
            skipped += 1
            continue

        item = ParsedItem(
            sap_code=sap_code,
            modelo=_coerce_text(_cell("modelo")),
            descripcion=_coerce_text(_cell("descripcion")),
            familia=_coerce_text(_cell("familia")),
            categoria=_coerce_text(_cell("categoria")),
            subcategoria=_coerce_text(_cell("subcategoria")),
            marca=_coerce_text(_cell("marca")),
            **prices,
        )
        if sap_code in items:
            duplicates += 1
        items[sap_code] = item

    return ParseResult(
        items=list(items.values()),
        duplicates_overwritten=duplicates,
        rows_skipped=skipped,
    )
