#!/usr/bin/env python3
"""Pre-procesa el Excel del cotizador a JSONs consumibles por el servicio shipping.

Uso:
    python3 scripts/build_shipping_data.py

Lee ``docs/Cotizador - Tarifa COM. DE ART. DE PROTECCION 2026_1S T.xlsx``
y genera dos archivos en ``src/services/shipping/data/``:

- ``localidades.json``: mapping ``localidad_normalizada -> {sucursal, zona,
  display_name}`` para los 510 destinos del Excel.
- ``tarifas.json``: índice ``[origen][destino][zona][tramo] -> {tarifa,
  minimo}`` con todas las rutas (incluye los 22 orígenes para no perder data
  el día que cambien la sucursal de despacho).

El script es idempotente: borra y reescribe los archivos cada vez. Re-ejecutar
cuando llegue una versión nueva del Excel.
"""
from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl  # type: ignore
except ImportError:
    print(
        "openpyxl no está instalado. Instalá con: pip install openpyxl",
        file=sys.stderr,
    )
    raise


def project_root() -> Path:
    return Path(__file__).resolve().parent.parent


def normalize_locality(name: str) -> str:
    """Mayúsculas, sin tildes, sin signos. Compara robusto contra inputs."""
    s = (name or "").strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def main() -> int:
    root = project_root()
    xlsx = (
        root
        / "docs"
        / "Cotizador - Tarifa COM. DE ART. DE PROTECCION 2026_1S T.xlsx"
    )
    if not xlsx.exists():
        print(f"error: no existe {xlsx}", file=sys.stderr)
        return 1

    print(f"Leyendo {xlsx.name} ...")
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # ----- LOCALIDADES -----
    if "LOCALIDADES" not in wb.sheetnames:
        print("error: hoja LOCALIDADES no existe", file=sys.stderr)
        return 1
    ws = wb["LOCALIDADES"]
    localidades: dict[str, dict] = {}
    duplicates: list[str] = []
    for r in range(2, ws.max_row + 1):
        loc = ws.cell(row=r, column=1).value
        suc = ws.cell(row=r, column=2).value
        zona = ws.cell(row=r, column=3).value
        if not loc or not suc or not zona:
            continue
        key = normalize_locality(str(loc))
        if not key:
            continue
        entry = {
            "sucursal": str(suc).strip().upper(),
            "zona": str(zona).strip().upper(),
            "display_name": str(loc).strip(),
        }
        if key in localidades and localidades[key] != entry:
            duplicates.append(key)
        localidades[key] = entry
    if duplicates:
        print(
            f"warning: {len(set(duplicates))} localidad(es) con datos diferentes "
            "se sobrescribieron con la última fila del Excel"
        )
    print(f"  localidades: {len(localidades)}")

    # ----- TARIFAS -----
    if "TARIFAS" not in wb.sheetnames:
        print("error: hoja TARIFAS no existe", file=sys.stderr)
        return 1
    ws = wb["TARIFAS"]
    tarifas: dict[str, dict[str, dict[str, dict[str, dict]]]] = {}
    rows = 0
    skipped = 0
    for r in range(5, ws.max_row + 1):
        ruta = ws.cell(row=r, column=2).value
        origen = ws.cell(row=r, column=3).value
        zona = ws.cell(row=r, column=4).value
        destino = ws.cell(row=r, column=5).value
        tramo = ws.cell(row=r, column=6).value
        tarifa = ws.cell(row=r, column=7).value
        minimo = ws.cell(row=r, column=8).value
        if not ruta:
            continue
        if origen is None or destino is None or zona is None or tramo is None:
            skipped += 1
            continue
        try:
            t_val = float(tarifa) if tarifa is not None else 0.0
            m_val = float(minimo) if minimo is not None else 0.0
        except (TypeError, ValueError):
            skipped += 1
            continue
        o = str(origen).strip().upper()
        d = str(destino).strip().upper()
        z = str(zona).strip().upper()
        t = str(tramo).strip()
        tarifas.setdefault(o, {}).setdefault(d, {}).setdefault(z, {})[t] = {
            "tarifa": round(t_val, 4),
            "minimo": round(m_val, 2),
        }
        rows += 1
    print(f"  tarifas: {rows} filas indexadas (skipped: {skipped})")

    # ----- TRAMOS (orden + bordes) -----
    # Inferir tramos del Excel para garantizar consistencia con la fuente.
    tramos_set: set[str] = set()
    for _o, dests in tarifas.items():
        for _d, zonas in dests.items():
            for _z, ts in zonas.items():
                tramos_set.update(ts.keys())

    def parse_tramo(t: str) -> tuple[float, float]:
        m = re.match(r"^\s*(\d+)\s*-\s*(\d+)\s*$", t)
        if not m:
            return (0.0, 0.0)
        return (float(m.group(1)), float(m.group(2)))

    tramos = sorted(tramos_set, key=lambda s: parse_tramo(s)[0])
    tramos_meta = [
        {"tramo": t, "from_kg": parse_tramo(t)[0], "to_kg": parse_tramo(t)[1]}
        for t in tramos
    ]
    print(f"  tramos: {len(tramos)} -> {tramos}")

    # ----- WRITE -----
    out_dir = root / "src" / "services" / "shipping" / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "localidades.json").write_text(
        json.dumps(localidades, ensure_ascii=False, sort_keys=True, indent=2),
        encoding="utf-8",
    )
    (out_dir / "tarifas.json").write_text(
        json.dumps(tarifas, ensure_ascii=False, sort_keys=True),
        encoding="utf-8",
    )
    (out_dir / "tramos.json").write_text(
        json.dumps(tramos_meta, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Listo: {out_dir.relative_to(root)}/{{localidades,tarifas,tramos}}.json")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
